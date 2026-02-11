from fastapi import APIRouter, Depends, Request, UploadFile, File, Form
from fastapi.responses import StreamingResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook
import io, json, openpyxl
from datetime import datetime
from .. import models, database, utils

router = APIRouter(tags=["Excel"])
templates = Jinja2Templates(directory="templates")

def escanear_excel_flexible(ws):
    datos = { "tipo": "EQUIPO", "marca": "", "modelo": "", "referencia": "", "serial": "", "ip": "", "mac": "", "hostname": "", "responsable": "", "estado": "Bueno", "accesorios": [] }
    fila_8 = (str(ws['B8'].value or "") + str(ws['C8'].value or "") + str(ws['D8'].value or "")).upper()
    if "PORTATIL" in fila_8: datos["tipo"] = "PORTATIL"
    elif "ESCRITORIO" in fila_8: datos["tipo"] = "COMPUTADOR"
    for r in range(5, 35):
        lbl = str(ws[f'B{r}'].value or "").upper(); val = str(ws[f'C{r}'].value or "").strip()
        if "MARCA" in lbl: datos["marca"] = val.upper()
        if "MODELO" in lbl: datos["modelo"] = val
        if "SERIAL" in lbl or "SERIE" in lbl: datos["serial"] = utils.limpiar_serial(val)
    return datos

@router.post("/prellenar_desde_excel")
async def prellenar_formulario(request: Request, file: UploadFile = File(...), db: Session = Depends(database.get_db)):
    if not request.session.get("usuario"): return RedirectResponse("/login")
    try:
        contents = await file.read()
        wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
        info = escanear_excel_flexible(wb.active)
        info["accesorios_json"] = json.dumps(info["accesorios"])
        return templates.TemplateResponse("crear.html", {
            "request": request, "pre_data": info, "tipos": db.query(models.TipoEquipo).all(), 
            "marcas": db.query(models.Marca).all(), "modelos": db.query(models.Modelo).all(), 
            "areas": db.query(models.Area).all(),
            "posibles_padres": db.query(models.ActivoTec).filter(models.ActivoTec.id_padre_activo == None).all()
        })
    except Exception as e: return RedirectResponse(f"/crear?msg=Error: {e}", status_code=303)
