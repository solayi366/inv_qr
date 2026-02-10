import uuid 
import csv
import io
import json
import re
import openpyxl
from datetime import datetime
from typing import List, Optional

from fastapi import FastAPI, Depends, HTTPException, Request, Form, UploadFile, File
from fastapi.templating import Jinja2Templates 
from fastapi.staticfiles import StaticFiles
from fastapi.responses import RedirectResponse, StreamingResponse
from starlette.middleware.sessions import SessionMiddleware 
from sqlalchemy.orm import Session
from sqlalchemy.exc import IntegrityError 
from sqlalchemy import func 
from passlib.context import CryptContext 

from openpyxl import Workbook, load_workbook

from . import models, schemas, database, utils

# --- CONFIGURACIÓN ---
BASE_URL = "https://inventario.envia06.com/" 
SECRET_KEY = "SUPXROSEVREUOIC4MBIPME" 

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
models.Base.metadata.create_all(bind=database.engine)

app = FastAPI(title="Sistema Inventario QR")
app.add_middleware(SessionMiddleware, secret_key=SECRET_KEY)
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates") 

def get_db():
    db = database.SessionLocal()
    try:
        yield db
    finally:
        db.close()

# --- FUNCIONES AUXILIARES ---
def buscar_o_crear(db: Session, modelo, **filtros):
    instancia = db.query(modelo).filter_by(**filtros).first()
    if instancia: return instancia
    else:
        nueva_instancia = modelo(**filtros)
        db.add(nueva_instancia); db.commit(); db.refresh(nueva_instancia)
        return nueva_instancia

def limpiar_serial(valor):
    if not valor: return None
    s = str(valor).strip().upper()
    BLACKLIST = ["N.A", "NA", "NONE", "SERIAL", "SERIE", "MARCA", "MODELO", "REFERENCIA", "GENERICO", "", "MOUSE", "TECLADO", "NULL"]
    if s in BLACKLIST or len(s) < 3: return None
    return s.split('/')[0].strip()

def limpiar_mac(valor):
    if not valor: return None
    s = str(valor).strip().upper().replace(" ", "")
    return s[:17]

def get_password_hash(password): return pwd_context.hash(password)
def verify_password(plain, hashed): return pwd_context.verify(plain, hashed)

# --- FUNCIÓN DE AUDITORÍA ---
def registrar_historia(db: Session, id_activo: int, tipo: str, descripcion: str, usuario: str):
    try:
        nuevo = models.Actualizacion(
            id_activo=id_activo,
            tipo_evento=tipo,
            desc_evento=descripcion[:250], 
            usuario_sistema=usuario,
            fecha=datetime.now()
        )
        db.add(nuevo)
    except Exception as e:
        print(f"Error guardando historial: {e}")

# --- API ENDPOINTS (AJAX/FETCH) ---
@app.get("/api/empleado/{id}")
def api_buscar_empleado(id: str, db: Session = Depends(get_db)):
    id_clean = id.strip()
    empleado = db.query(models.Empleado).filter(models.Empleado.cod_nom == id_clean).first()
    
    if not empleado:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")
    
    return {
        "cod_nom": empleado.cod_nom,
        "nombre": empleado.nom_emple,
        "id_area": empleado.id_area,
        "nom_area": empleado.area.nom_area if empleado.area else "SIN ÁREA"
    }

# --- API DE ESTADÍSTICAS (ACTUALIZADA) ---
@app.get("/api/stats")
def api_estadisticas(request: Request, db: Session = Depends(get_db)):
    if not request.session.get("usuario"): raise HTTPException(status_code=401)
    
    # 1. Totales Generales
    total_activos = db.query(models.ActivoTec).count()
    
    # 2. Tickets Pendientes (NUEVO)
    tickets_pendientes = db.query(models.Novedad).filter(
        models.Novedad.estado_ticket != "Cerrado"
    ).count()
    
    # 3. Por Estado
    por_estado = db.query(models.ActivoTec.estado, func.count(models.ActivoTec.id_activo))\
                   .group_by(models.ActivoTec.estado).all()
    
    # 4. Por Marca (Top 5)
    por_marca = db.query(models.Marca.nom_marca, func.count(models.ActivoTec.id_activo))\
                  .join(models.Marca)\
                  .group_by(models.Marca.nom_marca)\
                  .order_by(func.count(models.ActivoTec.id_activo).desc())\
                  .limit(5).all()

    # 5. Por Tipo
    por_tipo = db.query(models.TipoEquipo.nom_tipo, func.count(models.ActivoTec.id_activo))\
                 .join(models.TipoEquipo)\
                 .group_by(models.TipoEquipo.nom_tipo).all()

    return {
        "total": total_activos,
        "pendientes": tickets_pendientes, # <-- Enviamos este dato nuevo
        "estados": [{"label": e[0], "count": e[1]} for e in por_estado],
        "marcas": [{"label": m[0], "count": m[1]} for m in por_marca],
        "tipos": [{"label": t[0], "count": t[1]} for t in por_tipo]
    }

# --- LOGIN Y SISTEMA ---
@app.on_event("startup")
def startup_event():
    db = database.SessionLocal()
    user = db.query(models.Usuario).filter(models.Usuario.username == "admin").first()
    if not user:
        db.add(models.Usuario(username="admin", contraseña=get_password_hash("admin123")))
        db.commit()
    db.close()

@app.get("/login", response_class=templates.TemplateResponse)
def login_form(request: Request): return templates.TemplateResponse("login.html", {"request": request})

@app.post("/login")
def login(request: Request, username: str = Form(...), password: str = Form(...), db: Session = Depends(get_db)):
    user = db.query(models.Usuario).filter(models.Usuario.username == username).first()
    if not user or not verify_password(password, user.contraseña):
        return templates.TemplateResponse("index.html", {"request": request, "error": "Credenciales inválidas"})
    request.session["usuario"] = user.username
    return RedirectResponse(url="/?msg=Bienvenido al Sistema&tipo=success", status_code=303)

@app.get("/logout")
def logout(request: Request):
    request.session.clear()
    return RedirectResponse(url="/login?msg=Sesión cerrada&tipo=info", status_code=303)

@app.get("/", response_class=templates.TemplateResponse)
def pagina_principal(request: Request, db: Session = Depends(get_db)):
    return templates.TemplateResponse("index.html", {
        "request": request, "activos": db.query(models.ActivoTec).all(), "user": request.session.get("usuario")
    })

@app.get("/dashboard", response_class=templates.TemplateResponse)
def dashboard_view(request: Request):
    if not request.session.get("usuario"): return RedirectResponse("/login", status_code=303)
    return templates.TemplateResponse("dashboard.html", {"request": request})

@app.get("/ver/{id}")
def ver_hoja_vida(request: Request, id: int, db: Session = Depends(get_db)):
    activo = db.query(models.ActivoTec).filter(models.ActivoTec.id_activo == id).first()
    if not activo: raise HTTPException(status_code=404)
    return templates.TemplateResponse("ficha.html", {
        "request": request, 
        "activo": activo, 
        "user": request.session.get("usuario")
    })

@app.get("/historial/{id}")
def ver_historial_activo(request: Request, id: int, db: Session = Depends(get_db)):
    if not request.session.get("usuario"): return RedirectResponse("/login", status_code=303)
    
    activo = db.query(models.ActivoTec).filter(models.ActivoTec.id_activo == id).first()
    if not activo: raise HTTPException(status_code=404)
    
    historial = db.query(models.Actualizacion).filter(
        models.Actualizacion.id_activo == id
    ).order_by(models.Actualizacion.fecha.desc()).all()
    
    return templates.TemplateResponse("historial.html", {
        "request": request, 
        "activo": activo,
        "historial": historial
    })

@app.get("/imprimir/{id}", response_class=templates.TemplateResponse)
def imprimir_etiqueta(request: Request, id: int, db: Session = Depends(get_db)):
    if not request.session.get("usuario"): return RedirectResponse("/login", status_code=303)
    activo = db.query(models.ActivoTec).filter(models.ActivoTec.id_activo == id).first()
    if not activo: raise HTTPException(status_code=404)
    return templates.TemplateResponse("imprimir.html", {"request": request, "activo": activo})

@app.get("/exportar")
def exportar_inventario(request: Request, db: Session = Depends(get_db)):
    if not request.session.get("usuario"): return RedirectResponse("/login", status_code=303)
    
    wb = Workbook(); ws = wb.active; ws.title = "Inventario Completo"
    headers = ["ID", "Tipo", "Marca", "Modelo", "Serial", "Referencia", "Hostname", "IP", "MAC", "Estado", "Responsable", "Área", "Ubicación"]
    ws.append(headers)
    for cell in ws[1]: cell.font = openpyxl.styles.Font(bold=True)
    
    activos = db.query(models.ActivoTec).all()
    for a in activos:
        nom_modelo = a.modelo.nom_modelo if a.modelo else ""
        if a.responsable:
            nom_resp = a.responsable.nom_emple; nom_area = a.responsable.area.nom_area if a.responsable.area else ""
            ubicacion = "Asignado"
        else:
            nom_resp = "SIN ASIGNAR"; nom_area = "-"; ubicacion = "BODEGA"

        ws.append([a.id_activo, a.tipo.nom_tipo, a.marca.nom_marca, nom_modelo, a.serial, a.referencia or "", a.hostname or "", a.ip_equipo or "", a.mac_activo or "", a.estado, nom_resp, nom_area, ubicacion])

    for col in ws.columns:
        max_length = 0; column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = (max_length + 2)

    output = io.BytesIO(); wb.save(output); output.seek(0)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=Inventario_{timestamp}.xlsx"})

@app.get("/exportar_individual/{id}")
def exportar_individual(request: Request, id: int, db: Session = Depends(get_db)):
    if not request.session.get("usuario"): return RedirectResponse("/login", status_code=303)
    
    activo = db.query(models.ActivoTec).filter(models.ActivoTec.id_activo == id).first()
    if not activo: raise HTTPException(status_code=404)

    wb = Workbook(); ws = wb.active; ws.title = f"Activo {activo.id_activo}"
    bold_font = openpyxl.styles.Font(bold=True)
    header_fill = openpyxl.styles.PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    center_align = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    
    ws.merge_cells('A1:B1'); ws['A1'] = f"HOJA DE VIDA - ACTIVO #{activo.codigo_qr}"
    ws['A1'].font = openpyxl.styles.Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].alignment = center_align; ws['A1'].fill = openpyxl.styles.PatternFill(start_color="B33515", end_color="B33515", fill_type="solid")

    datos = [("Tipo", activo.tipo.nom_tipo), ("Marca", activo.marca.nom_marca), ("Modelo", activo.modelo.nom_modelo if activo.modelo else "Genérico"), ("Serial", activo.serial), ("Referencia", activo.referencia or "N/A"), ("Estado", activo.estado), ("QR", activo.codigo_qr)]
    row = 3
    for label, value in datos:
        c = ws.cell(row=row, column=1, value=label); c.font = bold_font; c.fill = header_fill
        ws.cell(row=row, column=2, value=value); row += 1

    if activo.hostname or activo.ip_equipo or activo.mac_activo:
        row += 1; ws.merge_cells(f'A{row}:B{row}'); ws[f'A{row}'] = "CONECTIVIDAD"; ws[f'A{row}'].font = bold_font; ws[f'A{row}'].alignment = center_align; row += 1
        red = [("Hostname", activo.hostname), ("IP", activo.ip_equipo), ("MAC", activo.mac_activo)]
        for l, v in red: ws.cell(row=row, column=1, value=l).font = bold_font; ws.cell(row=row, column=2, value=v or "-"); row += 1

    row += 1; ws.merge_cells(f'A{row}:B{row}'); ws[f'A{row}'] = "ASIGNACIÓN"; ws[f'A{row}'].font = bold_font; ws[f'A{row}'].alignment = center_align; row += 1
    if activo.responsable:
        ws.cell(row=row, column=1, value="Responsable").font = bold_font; ws.cell(row=row, column=2, value=activo.responsable.nom_emple); row += 1
        ws.cell(row=row, column=1, value="ID").font = bold_font; ws.cell(row=row, column=2, value=activo.responsable.cod_nom); row += 1
        ws.cell(row=row, column=1, value="Área").font = bold_font; ws.cell(row=row, column=2, value=activo.responsable.area.nom_area if activo.responsable.area else "-")
    else:
        ws.cell(row=row, column=1, value="Ubicación").font = bold_font; ws.cell(row=row, column=2, value="En Bodega")
    
    if activo.hijos:
        row += 2; ws.merge_cells(f'A{row}:D{row}'); ws[f'A{row}'] = "ACCESORIOS"; ws[f'A{row}'].font = bold_font; ws[f'A{row}'].fill = header_fill; row += 1
        for i, h in enumerate(["ID", "Tipo", "Marca", "Serial"], 1): ws.cell(row=row, column=i, value=h).font = bold_font
        row += 1
        for h in activo.hijos:
            ws.cell(row=row, column=1, value=h.codigo_qr); ws.cell(row=row, column=2, value=h.tipo.nom_tipo)
            ws.cell(row=row, column=3, value=h.marca.nom_marca); ws.cell(row=row, column=4, value=h.serial); row += 1

    ws.column_dimensions['A'].width = 25; ws.column_dimensions['B'].width = 40; ws.column_dimensions['C'].width = 20; ws.column_dimensions['D'].width = 25
    output = io.BytesIO(); wb.save(output); output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=HojaVida_{activo.serial}.xlsx"})

def escanear_excel_flexible(ws):
    datos = { "tipo": "EQUIPO", "marca": "", "modelo": "", "referencia": "", "serial": "", "ip": "", "mac": "", "hostname": "", "responsable": "", "estado": "Bueno", "accesorios": [] }
    fila_8 = (str(ws['B8'].value or "") + str(ws['C8'].value or "") + str(ws['D8'].value or "")).upper()
    val_b10 = str(ws['B10'].value or "").strip().upper()
    
    if "PORTATIL" in fila_8 and ("_X_" in fila_8 or " X " in fila_8 or "X" in str(ws['B8'].value or "").upper()): datos["tipo"] = "PORTATIL"
    elif "ESCRITORIO" in fila_8 and ("_X_" in fila_8 or " X " in fila_8): datos["tipo"] = "COMPUTADOR"
    elif "TODO EN UNO" in fila_8 and ("_X_" in fila_8 or " X " in fila_8): datos["tipo"] = "AIO"
    elif val_b10 in ["TABLET", "PORTATIL"]: datos["tipo"] = val_b10
    elif val_b10 == "CPU" or "CPU" in str(ws['B11'].value or "").upper(): datos["tipo"] = "COMPUTADOR"

    for r in range(5, 35):
        label = str(ws[f'B{r}'].value or "").strip().upper(); val = str(ws[f'C{r}'].value or "").strip()
        if not label: continue
        if "NOMBRE EQUIPO" in label: datos["hostname"] = val
        elif label == "MARCA": datos["marca"] = val.upper()
        elif label == "MODELO": datos["modelo"] = val
        elif label == "REFERENCIA": datos["referencia"] = val
        elif label in ["SERIAL", "SERIE"]: 
            s = limpiar_serial(val)
            if s: datos["serial"] = s
        elif label == "IP": datos["ip"] = val
        elif label in ["MAC", "MAC:", "DIRECCION MAC"]: datos["mac"] = limpiar_mac(val)
        elif "USUARIO:" in label: datos["responsable"] = val
        elif "RESPONSABLE" in label: datos["responsable"] = val if len(val) > 2 else str(ws[f'B{r + 1}'].value or "").strip()

    if not datos["serial"]:
        c13 = limpiar_serial(ws['C13'].value); c14 = limpiar_serial(ws['C14'].value)
        if c13: datos["serial"] = c13
        elif c14: datos["serial"] = c14

    keywords_acc = ["TECLADO", "MOUSE", "MONITOR", "LECTOR", "UPS", "BASE", "CAMARA", "TARJETA"]
    def buscar_accesorio(fila, col_lbl, col_val):
        tipo = str(ws[f'{col_lbl}{fila}'].value or "").strip().upper()
        if not any(k in tipo for k in keywords_acc): return
        acc_marca = "GENERICO"; acc_serie = None; acc_ref = None
        for offset in range(1, 6):
            lbl = str(ws[f'{col_lbl}{fila+offset}'].value or "").upper(); val = str(ws[f'{col_val}{fila+offset}'].value or "").strip()
            if not lbl: break
            if "MARCA" in lbl: acc_marca = val.upper()
            if "REFERENCIA" in lbl: acc_ref = val
            if "SERIE" in lbl and "ENVIA" not in lbl: acc_serie = limpiar_serial(val)
        if acc_serie: datos["accesorios"].append({"tipo": tipo, "marca": acc_marca, "serial": acc_serie, "referencia": acc_ref or ""})

    for r in range(10, 45): buscar_accesorio(r, 'B', 'C'); buscar_accesorio(r, 'E', 'F')
    return datos

@app.post("/prellenar_desde_excel")
async def prellenar_formulario(request: Request, file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not request.session.get("usuario"): return RedirectResponse("/login", status_code=303)
    try:
        contents = await file.read(); wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
        info = escanear_excel_flexible(wb.active)
        info["accesorios_json"] = json.dumps(info["accesorios"])
        msg_tipo = "success" if info["serial"] else "warning"
        msg = f"Datos cargados. {len(info['accesorios'])} accesorios detectados."
        return templates.TemplateResponse("crear.html", {
            "request": request, "pre_data": info, "msg": msg, "tipo_msg": msg_tipo,
            "tipos": db.query(models.TipoEquipo).all(), "marcas": db.query(models.Marca).all(),
            "modelos": db.query(models.Modelo).all(), "areas": db.query(models.Area).all(),
            "posibles_padres": db.query(models.ActivoTec).filter(models.ActivoTec.id_padre_activo == None).all()
        })
    except Exception as e: return RedirectResponse(f"/crear?msg=Error leyendo archivo: {str(e)}&tipo=danger", status_code=303)

@app.get("/crear", response_class=templates.TemplateResponse)
def form_crear(request: Request, db: Session = Depends(get_db)):
    if not request.session.get("usuario"): return RedirectResponse("/login", status_code=303)
    return templates.TemplateResponse("crear.html", {
        "request": request, "tipos": db.query(models.TipoEquipo).all(), "marcas": db.query(models.Marca).all(), 
        "modelos": db.query(models.Modelo).all(), "areas": db.query(models.Area).all(),
        "posibles_padres": db.query(models.ActivoTec).filter(models.ActivoTec.id_padre_activo == None).all()
    })

@app.post("/crear")
def procesar_crear(request: Request, serial: str = Form(...), id_tipoequi: int = Form(...), id_marca: int = Form(...), estado: str = Form(...), 
                   referencia: str = Form(None), hostname: str = Form(None), ip_equipo: str = Form(None), mac_activo: str = Form(None), 
                   id_modelo: int = Form(None), id_padre_activo: int = Form(None), cod_responsable: str = Form(None), 
                   nom_nuevo_empleado: str = Form(None), id_area_nuevo: int = Form(None), 
                   accesorios_json_final: str = Form(None),
                   db: Session = Depends(get_db)):
    
    usuario_actual = request.session.get("usuario")
    if not usuario_actual: return RedirectResponse("/login", status_code=303)
    
    try:
        responsable_final = None
        if cod_responsable and cod_responsable.strip():
            emp = db.query(models.Empleado).filter(models.Empleado.cod_nom == cod_responsable.strip()).first()
            if emp: responsable_final = emp.cod_nom
            elif nom_nuevo_empleado:
                nuevo_id = cod_responsable.strip() if any(char.isdigit() for char in cod_responsable) else "T-" + str(uuid.uuid4().hex[:4]).upper()
                area_id = id_area_nuevo if id_area_nuevo else db.query(models.Area).first().id_area
                nuevo_emp = models.Empleado(cod_nom=nuevo_id, nom_emple=(nom_nuevo_empleado or cod_responsable).upper(), id_area=area_id, activo=True)
                db.add(nuevo_emp); db.commit(); responsable_final = nuevo_emp.cod_nom

        mac_final = limpiar_mac(mac_activo)
        nuevo = models.ActivoTec(
            serial=serial, referencia=referencia, hostname=hostname, ip_equipo=ip_equipo, mac_activo=mac_final,
            id_tipoequi=id_tipoequi, id_marca=id_marca, id_modelo=id_modelo if id_modelo and id_modelo>0 else None,
            estado=estado, id_padre_activo=id_padre_activo, cod_nom_responsable=responsable_final, codigo_qr=str(uuid.uuid4())
        )
        db.add(nuevo); db.commit(); db.refresh(nuevo)
        
        url_real = str(request.url_for('ver_hoja_vida', id=nuevo.id_activo))
        nuevo.codigo_qr = f"ACT-{nuevo.id_activo:04d}"
        utils.generar_codigo_qr(nuevo.codigo_qr, url_real)
        
        registrar_historia(db, nuevo.id_activo, "CREACION", f"Activo creado. Resp: {responsable_final or 'Bodega'}", usuario_actual)
        db.commit()

        if accesorios_json_final:
            try:
                lista_acc = json.loads(accesorios_json_final)
                for acc in lista_acc:
                    t_h = buscar_o_crear(db, models.TipoEquipo, nom_tipo=acc['tipo'])
                    m_h = buscar_o_crear(db, models.Marca, nom_marca=acc['marca'])
                    hijo = models.ActivoTec(
                        serial=acc['serial'], referencia=acc.get('referencia'), 
                        id_tipoequi=t_h.id_tipoequi, id_marca=m_h.id_marca, 
                        estado="Bueno", id_padre_activo=nuevo.id_activo, codigo_qr=str(uuid.uuid4())
                    )
                    db.add(hijo); db.commit(); db.refresh(hijo)
                    url_hijo = str(request.url_for('ver_hoja_vida', id=hijo.id_activo))
                    hijo.codigo_qr = f"ACC-{hijo.id_activo:04d}"
                    utils.generar_codigo_qr(hijo.codigo_qr, url_hijo)
                    
                    registrar_historia(db, hijo.id_activo, "CREACION", f"Accesorio vinculado a {nuevo.codigo_qr}", usuario_actual)
                    db.commit()
            except Exception: pass

        return RedirectResponse(f"/ver/{nuevo.id_activo}?msg=Activo Creado&tipo=success", status_code=303)
    except Exception as e: return RedirectResponse(f"/?msg=Error: {str(e)}&tipo=danger", status_code=303)

@app.get("/editar/{id}", response_class=templates.TemplateResponse)
def form_editar(request: Request, id: int, db: Session = Depends(get_db)):
    if not request.session.get("usuario"): return RedirectResponse("/login", status_code=303)
    activo = db.query(models.ActivoTec).filter(models.ActivoTec.id_activo == id).first()
    return templates.TemplateResponse("editar.html", {
        "request": request, "activo": activo, "tipos": db.query(models.TipoEquipo).all(), "marcas": db.query(models.Marca).all(),
        "modelos": db.query(models.Modelo).all(), "areas": db.query(models.Area).all(),
        "posibles_padres": db.query(models.ActivoTec).filter(models.ActivoTec.id_activo != id, models.ActivoTec.id_padre_activo == None).all()
    })

@app.post("/editar/{id}")
def procesar_editar(request: Request, id: int, serial: str = Form(...), id_tipoequi: int = Form(...), id_marca: int = Form(...), estado: str = Form(...),
                    referencia: str = Form(None), ip_equipo: str = Form(None), mac_activo: str = Form(None), id_modelo: int = Form(None),
                    id_padre_activo: int = Form(None), cod_responsable: str = Form(None), db: Session = Depends(get_db)):
    
    usuario_actual = request.session.get("usuario")
    if not usuario_actual: return RedirectResponse("/login", status_code=303)
    
    try:
        a = db.query(models.ActivoTec).get(id)
        
        cambios = []
        
        old_ser = (a.serial or "").strip()
        new_ser = (serial or "").strip()
        if old_ser != new_ser: cambios.append(f"Serial: {old_ser}->{new_ser}")

        if a.id_tipoequi != id_tipoequi:
            old_tipo = a.tipo.nom_tipo if a.tipo else "N/A"
            new_tipo_obj = db.query(models.TipoEquipo).get(id_tipoequi)
            new_tipo = new_tipo_obj.nom_tipo if new_tipo_obj else "N/A"
            cambios.append(f"Tipo: {old_tipo}->{new_tipo}")

        if a.id_marca != id_marca:
            old_marca = a.marca.nom_marca if a.marca else "N/A"
            new_marca_obj = db.query(models.Marca).get(id_marca)
            new_marca = new_marca_obj.nom_marca if new_marca_obj else "N/A"
            cambios.append(f"Marca: {old_marca}->{new_marca}")

        old_mod_id = a.id_modelo or 0
        new_mod_id = id_modelo or 0
        if old_mod_id != new_mod_id:
            cambios.append(f"Modelo modificado")

        old_ref = (a.referencia or "").strip()
        new_ref = (referencia or "").strip()
        if old_ref != new_ref: cambios.append(f"Ref: '{old_ref}' -> '{new_ref}'")

        if a.estado != estado: cambios.append(f"Estado: {a.estado}->{estado}")
        if (a.ip_equipo or "") != (ip_equipo or ""): cambios.append(f"IP: {a.ip_equipo}->{ip_equipo}")
        if (a.mac_activo or "") != (mac_activo or ""): cambios.append(f"MAC: {a.mac_activo}->{mac_activo}")
        
        nuevo_resp = None
        if cod_responsable and cod_responsable.strip():
            emp = db.query(models.Empleado).filter(models.Empleado.cod_nom == cod_responsable.strip()).first()
            if emp: nuevo_resp = emp.cod_nom
        
        old_resp = a.cod_nom_responsable or "Bodega"
        new_resp_val = nuevo_resp or "Bodega"
        
        if old_resp != new_resp_val:
            cambios.append(f"Responsable: {old_resp} -> {new_resp_val}")

        if cambios:
            registrar_historia(db, a.id_activo, "EDICION", "; ".join(cambios), usuario_actual)

        a.serial = serial; a.referencia = referencia; a.ip_equipo = ip_equipo; a.mac_activo = mac_activo
        a.id_tipoequi = id_tipoequi; a.id_marca = id_marca; a.estado = estado; a.id_padre_activo = id_padre_activo
        a.id_modelo = id_modelo if id_modelo and id_modelo > 0 else None
        a.cod_nom_responsable = nuevo_resp

        db.commit()
        return RedirectResponse(f"/ver/{id}?msg=Cambios Guardados&tipo=success", status_code=303)
    except Exception as e:
        print(f"Error al editar: {e}")
        return RedirectResponse(f"/ver/{id}?msg=Error al editar&tipo=danger", status_code=303)
        
# --- RUTA DE ELIMINACIÓN ---
@app.get("/eliminar/{id}")
def eliminar_activo(request: Request, id: int, db: Session = Depends(get_db)):
    if not request.session.get("usuario"): return RedirectResponse("/login", status_code=303)
    try:
        activo = db.query(models.ActivoTec).get(id)
        if not activo: return RedirectResponse("/?msg=Activo no encontrado&tipo=danger", status_code=303)
        
        db.query(models.Actualizacion).filter(models.Actualizacion.id_activo == id).delete()
        hijos = db.query(models.ActivoTec).filter(models.ActivoTec.id_padre_activo == id).all()
        for h in hijos:
            db.query(models.Actualizacion).filter(models.Actualizacion.id_activo == h.id_activo).delete()
            db.delete(h)
        db.delete(activo)
        db.commit()
        return RedirectResponse("/?msg=Activo y accesorios eliminados correctamente&tipo=success", status_code=303)
    except Exception as e:
        db.rollback()
        return RedirectResponse(f"/?msg=Error al eliminar: {str(e)}&tipo=danger", status_code=303)

@app.get("/parametros", response_class=templates.TemplateResponse)
def gestionar_parametros(request: Request, db: Session = Depends(get_db)):
    if not request.session.get("usuario"): return RedirectResponse("/login", status_code=303)
    return templates.TemplateResponse("parametros.html", {
        "request": request, "marcas": db.query(models.Marca).all(), "modelos": db.query(models.Modelo).all(),
        "tipos": db.query(models.TipoEquipo).all(), "areas": db.query(models.Area).all(), "empleados": db.query(models.Empleado).all()
    })

@app.post("/{tipo}/crear")
def crear_parametro_generico(request: Request, tipo: str, db: Session = Depends(get_db),
                             nom_marca: str = Form(None), nom_tipo: str = Form(None), nom_area: str = Form(None),
                             nom_modelo: str = Form(None), id_marca: int = Form(None), id_tipoequi: int = Form(None),
                             cod_nom: str = Form(None), nom_emple: str = Form(None), id_area: int = Form(None)):
    if not request.session.get("usuario"): return RedirectResponse("/login", status_code=303)
    try:
        if tipo == "marca": db.add(models.Marca(nom_marca=nom_marca.upper()))
        elif tipo == "tipo": db.add(models.TipoEquipo(nom_tipo=nom_tipo.upper()))
        elif tipo == "area": db.add(models.Area(nom_area=nom_area))
        elif tipo == "modelo": db.add(models.Modelo(nom_modelo=nom_modelo, id_marca=id_marca, id_tipoequi=id_tipoequi))
        elif tipo == "empleado": db.add(models.Empleado(cod_nom=cod_nom, nom_emple=nom_emple.upper(), id_area=id_area, activo=True))
        db.commit()
        return RedirectResponse(f"/parametros?msg=Registro creado&tipo=success", status_code=303)
    except Exception as e: return RedirectResponse(f"/parametros?msg=Error: {str(e)}&tipo=danger", status_code=303)

@app.get("/{tipo}/eliminar/{id}")
def eliminar_parametro_generico(request: Request, tipo: str, id: str, db: Session = Depends(get_db)):
    if not request.session.get("usuario"): return RedirectResponse("/login", status_code=303)
    try:
        item = None
        if tipo == "marca": item = db.query(models.Marca).filter(models.Marca.id_marca == int(id)).first()
        elif tipo == "tipo": item = db.query(models.TipoEquipo).filter(models.TipoEquipo.id_tipoequi == int(id)).first()
        elif tipo == "area": item = db.query(models.Area).filter(models.Area.id_area == int(id)).first()
        elif tipo == "modelo": item = db.query(models.Modelo).filter(models.Modelo.id_modelo == int(id)).first()
        elif tipo == "empleado": item = db.query(models.Empleado).filter(models.Empleado.cod_nom == id).first()
        if item: db.delete(item); db.commit(); return RedirectResponse(f"/parametros?msg=Eliminado correctamente&tipo=success", status_code=303)
        return RedirectResponse(f"/parametros?msg=No encontrado&tipo=warning", status_code=303)
    except IntegrityError: db.rollback(); return RedirectResponse(f"/parametros?msg=No se puede eliminar: Está siendo usado&tipo=danger", status_code=303)

@app.get("/editar_parametro/{tipo}/{id}", response_class=templates.TemplateResponse)
def vista_editar_parametro(request: Request, tipo: str, id: str, db: Session = Depends(get_db)):
    if not request.session.get("usuario"): return RedirectResponse("/login", status_code=303)
    item = None
    if tipo == "marca": item = db.query(models.Marca).filter(models.Marca.id_marca == int(id)).first()
    elif tipo == "modelo": item = db.query(models.Modelo).filter(models.Modelo.id_modelo == int(id)).first()
    elif tipo == "tipo": item = db.query(models.TipoEquipo).filter(models.TipoEquipo.id_tipoequi == int(id)).first()
    elif tipo == "area": item = db.query(models.Area).filter(models.Area.id_area == int(id)).first()
    elif tipo == "empleado": item = db.query(models.Empleado).filter(models.Empleado.cod_nom == id).first()
    return templates.TemplateResponse("editar_parametro.html", {
        "request": request, "item": item, "tipo": tipo, "marcas": db.query(models.Marca).all(), 
        "tipos": db.query(models.TipoEquipo).all(), "areas": db.query(models.Area).all()
    })

@app.post("/editar_parametro/{tipo}/{id}")
def procesar_editar_parametro(request: Request, tipo: str, id: str, nombre: str = Form(...), nuevo_codigo: str = Form(None), id_marca: int = Form(None), id_tipoequi: int = Form(None), id_area: int = Form(None), db: Session = Depends(get_db)):
    if not request.session.get("usuario"): return RedirectResponse("/login", status_code=303)
    try:
        if tipo == "marca": db.query(models.Marca).filter(models.Marca.id_marca == int(id)).update({"nom_marca": nombre.upper()})
        elif tipo == "modelo": db.query(models.Modelo).filter(models.Modelo.id_modelo == int(id)).update({"nom_modelo": nombre, "id_marca": id_marca, "id_tipoequi": id_tipoequi})
        elif tipo == "tipo": db.query(models.TipoEquipo).filter(models.TipoEquipo.id_tipoequi == int(id)).update({"nom_tipo": nombre.upper()})
        elif tipo == "area": db.query(models.Area).filter(models.Area.id_area == int(id)).update({"nom_area": nombre})
        elif tipo == "empleado":
            codigo_actual = id
            codigo_nuevo = nuevo_codigo.strip() if nuevo_codigo else codigo_actual
            if codigo_nuevo != codigo_actual:
                if db.query(models.Empleado).filter(models.Empleado.cod_nom == codigo_nuevo).first(): return RedirectResponse(f"/parametros?msg=Error: La cédula {codigo_nuevo} ya existe&tipo=danger", status_code=303)
                nuevo_emp = models.Empleado(cod_nom=codigo_nuevo, nom_emple=nombre.upper(), id_area=id_area, activo=True)
                db.add(nuevo_emp); db.commit()
                activos = db.query(models.ActivoTec).filter(models.ActivoTec.cod_nom_responsable == codigo_actual).all()
                for a in activos: a.cod_nom_responsable = codigo_nuevo
                db.commit()
                db.query(models.Empleado).filter(models.Empleado.cod_nom == codigo_actual).delete()
            else:
                db.query(models.Empleado).filter(models.Empleado.cod_nom == codigo_actual).update({"nom_emple": nombre.upper(), "id_area": id_area})
        db.commit()
        return RedirectResponse(url="/parametros?msg=Editado correctamente&tipo=success", status_code=303)
    except Exception as e:
        db.rollback()
        return RedirectResponse(f"/parametros?msg=Error al editar: {str(e)}&tipo=danger", status_code=303)
    
    # --- MÓDULO DE NOVEDADES Y MANTENIMIENTO ---

@app.get("/novedad/{id}", response_class=templates.TemplateResponse)
def form_novedad(request: Request, id: int, db: Session = Depends(get_db)):
    """Muestra el formulario para reportar una novedad o mantenimiento"""
    if not request.session.get("usuario"): return RedirectResponse("/login", status_code=303)
    
    activo = db.query(models.ActivoTec).filter(models.ActivoTec.id_activo == id).first()
    if not activo: raise HTTPException(status_code=404)
    
    return templates.TemplateResponse("novedad.html", {
        "request": request, 
        "activo": activo
    })

@app.post("/novedad/{id}")
def registrar_novedad_evento(request: Request, id: int, tipo_novedad: str = Form(...), descripcion: str = Form(...), db: Session = Depends(get_db)):
    """Guarda la novedad en el historial del activo"""
    usuario_actual = request.session.get("usuario")
    if not usuario_actual: return RedirectResponse("/login", status_code=303)
    
    try:
        # Usamos la función existente registrar_historia pero con el tipo de novedad seleccionado
        # Esto nos ahorra crear tablas nuevas y centraliza todo en el timeline
        registrar_historia(db, id, tipo_novedad.upper(), descripcion, usuario_actual)
        db.commit()
        
        return RedirectResponse(f"/historial/{id}?msg=Novedad registrada correctamente&tipo=success", status_code=303)
    except Exception as e:
        print(f"Error novedad: {e}")
        return RedirectResponse(f"/ver/{id}?msg=Error al registrar: {str(e)}&tipo=danger", status_code=303)

@app.get("/sistema/reset-fabrica")
def reset_base_de_datos(request: Request, db: Session = Depends(get_db)):
    if not request.session.get("usuario"): return RedirectResponse("/login", status_code=303)
    try:
        db.query(models.ActivoTec).delete()
        db.query(models.Modelo).delete()
        db.query(models.Empleado).delete()
        db.query(models.Marca).delete()
        db.query(models.TipoEquipo).delete()
        db.query(models.Area).delete()
        db.commit()
        return RedirectResponse(f"/?msg=Sistema formateado (Admin conservado)&tipo=warning", status_code=303)
    except Exception as e:
        db.rollback()
        return RedirectResponse(f"/?msg=Error al limpiar: {str(e)}&tipo=danger", status_code=303)
    
# ==========================================
# MÓDULO DE REPORTE DE NOVEDADES (PORTAL)
# ==========================================

import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import shutil
import os

# Asegurar directorio de evidencias
os.makedirs("static/uploads", exist_ok=True)

# --- CONFIGURACIÓN CORREO (Ajusta esto) ---
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587
# ¡OJO! Si usas Gmail, la contraseña NO es la de tu login habitual.
# Debes generar una "Contraseña de Aplicación" en tu cuenta de Google > Seguridad.
SMTP_USER = "solayitapias1@gmail.com" 
SMTP_PASSWORD = "rrvz xzmd ngaw mxch" 
EMAIL_DESTINO = "sisbuca@envia.co"

def enviar_alerta_email(asunto, mensaje_html):
    print(f"📧 Intentando enviar correo a {EMAIL_DESTINO}...")
    try:
        msg = MIMEMultipart()
        msg['From'] = SMTP_USER
        msg['To'] = EMAIL_DESTINO
        msg['Subject'] = asunto
        msg.attach(MIMEText(mensaje_html, 'html'))

        # Conexión segura
        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        server.starttls() 
        server.login(SMTP_USER, SMTP_PASSWORD)
        text = msg.as_string()
        server.sendmail(SMTP_USER, EMAIL_DESTINO, text)
        server.quit()
        print("✅ Correo enviado EXITOSAMENTE.")
        return True
    except Exception as e:
        print(f"❌ ERROR CRÍTICO ENVIANDO CORREO: {str(e)}")
        return False

# 1. VISTA DEL PORTAL (Solo pide cédula al inicio)
@app.get("/portal-reportes", response_class=templates.TemplateResponse)
def vista_portal_reportes(request: Request):
    return templates.TemplateResponse("portal_reportes.html", {"request": request})

# 2. API: BUSCAR ACTIVOS POR CÉDULA (AJAX)
@app.get("/api/mis-activos/{cedula}")
def buscar_mis_activos(cedula: str, db: Session = Depends(get_db)):
    # Buscamos activos donde el responsable coincida
    # Nota: Asumimos que 'cod_nom_responsable' guarda la cédula o código
    activos = db.query(models.ActivoTec).filter(
        models.ActivoTec.cod_nom_responsable == cedula,
        models.ActivoTec.estado != "Baja" # No mostrar activos dados de baja
    ).all()
    
    if not activos:
        return {"encontrado": False, "mensaje": "No se encontraron activos asignados a este documento."}
    
    # Buscamos el nombre del empleado del primer activo para saludar
    nombre_empleado = activos[0].responsable.nom_emple if activos[0].responsable else "Usuario"

    lista_activos = []
    for a in activos:
        lista_activos.append({
            "id": a.id_activo,
            "tipo": a.tipo.nom_tipo,
            "marca": a.marca.nom_marca,
            "modelo": a.modelo.nom_modelo if a.modelo else "Genérico",
            "serial": a.serial,
            "foto_qr": f"/static/qrcodes/{a.codigo_qr}.png"
        })
        
    return {
        "encontrado": True, 
        "empleado": nombre_empleado,
        "activos": lista_activos
    }

# 3. PROCESAR EL REPORTE (Guarda en BD y envía correo)
@app.post("/crear-ticket")
async def crear_ticket_novedad(
    request: Request,
    cedula: str = Form(...),
    id_activo: int = Form(...),
    tipo_dano: str = Form(...),
    descripcion: str = Form(...),
    foto: UploadFile = File(None),
    db: Session = Depends(get_db)
):
    try:
        # 1. Procesar Foto (Si existe)
        ruta_foto = ""
        if foto and foto.filename:
            ext = foto.filename.split(".")[-1]
            nombre_archivo = f"ticket_{id_activo}_{uuid.uuid4().hex[:6]}.{ext}"
            ruta_guardada = f"static/uploads/{nombre_archivo}"
            with open(ruta_guardada, "wb") as buffer:
                shutil.copyfileobj(foto.file, buffer)
            ruta_foto = f"{BASE_URL}{ruta_guardada}"

        # 2. Buscar datos del activo para el correo y BD
        activo = db.query(models.ActivoTec).get(id_activo)
        nombre_empleado = activo.responsable.nom_emple if activo.responsable else "Desconocido"

        # 3. Guardar en Nueva Tabla Novedades
        nuevo_ticket = models.Novedad(
            cedula_reportante=cedula,
            nombre_reportante=nombre_empleado,
            id_activo=id_activo,
            tipo_daño=tipo_dano,
            descripcion=descripcion,
            evidencia_foto=ruta_foto
        )
        db.add(nuevo_ticket)
        
        # 4. Actualizar Estado del Activo (Opcional, pero recomendado)
        activo.estado = "Malo" # O "Reportado"
        
        # 5. Dejar rastro en el Historial General también
        registrar_historia(db, id_activo, "REPORTE_USUARIO", f"Falla reportada: {descripcion}", "PORTAL_WEB")
        
        db.commit()

        # 6. Enviar Correo
        asunto = f"🚨 TICKET #{nuevo_ticket.id_novedad} - {activo.tipo.nom_tipo} - {nombre_empleado}"
        html = f"""
        <div style="font-family: Arial, sans-serif; padding: 20px; border: 1px solid #ddd;">
            <h2 style="color: #d32f2f;">Nuevo Reporte de Falla</h2>
            <p>El usuario <strong>{nombre_empleado}</strong> (CC: {cedula}) ha reportado una novedad.</p>
            
            <table style="width: 100%; border-collapse: collapse; margin-top: 15px;">
                <tr style="background-color: #f9f9f9;"><td style="padding: 8px;"><strong>Activo:</strong></td><td style="padding: 8px;">{activo.tipo.nom_tipo} {activo.marca.nom_marca}</td></tr>
                <tr><td style="padding: 8px;"><strong>Serial:</strong></td><td style="padding: 8px;">{activo.serial}</td></tr>
                <tr style="background-color: #f9f9f9;"><td style="padding: 8px;"><strong>Tipo Daño:</strong></td><td style="padding: 8px;">{tipo_dano}</td></tr>
                <tr><td style="padding: 8px;"><strong>Descripción:</strong></td><td style="padding: 8px;">{descripcion}</td></tr>
            </table>
            
            <div style="margin-top: 20px;">
                <strong>Evidencia Fotográfica:</strong><br>
                {f'<img src="{ruta_foto}" width="300" style="border: 2px solid #333; margin-top:10px;">' if ruta_foto else '<i>No se adjuntó foto</i>'}
            </div>
            
            <p style="margin-top: 30px; font-size: 12px; color: #666;">Sistema Automático INVRQR</p>
        </div>
        """
        
        enviar_alerta_email(asunto, html)

        return templates.TemplateResponse("portal_exito.html", {"request": request, "id_ticket": nuevo_ticket.id_novedad})

    except Exception as e:
        print(f"Error creando ticket: {e}")
        return RedirectResponse("/portal-reportes?error=true", status_code=303)
    
    # --- GESTIÓN INTERNA DE NOVEDADES (ADMIN) ---

@app.get("/gestion-novedades", response_class=templates.TemplateResponse)
def panel_novedades(request: Request, db: Session = Depends(get_db)):
    """Bandeja de entrada de tickets no resueltos"""
    if not request.session.get("usuario"): return RedirectResponse("/login", status_code=303)
    
    # Traemos solo los que NO están cerrados, ordenados por fecha (más reciente primero)
    tickets = db.query(models.Novedad).filter(
        models.Novedad.estado_ticket != "Cerrado"
    ).order_by(models.Novedad.fecha_reporte.desc()).all()
    
    return templates.TemplateResponse("lista_novedades.html", {
        "request": request, 
        "tickets": tickets
    })

@app.post("/novedad/resolver/{id}")
def resolver_novedad(request: Request, id: int, solucion: str = Form(...), db: Session = Depends(get_db)):
    """Cierra el ticket y actualiza el historial"""
    if not request.session.get("usuario"): return RedirectResponse("/login", status_code=303)
    
    ticket = db.query(models.Novedad).get(id)
    if not ticket: return RedirectResponse("/gestion-novedades?msg=Ticket no encontrado&tipo=danger", status_code=303)
    
    # 1. Cambiar estado del ticket
    ticket.estado_ticket = "Cerrado"
    
    # 2. Dejar constancia en el historial del activo
    usuario = request.session.get("usuario")
    desc_cierre = f"TICKET #{id} CERRADO. Solución: {solucion}"
    registrar_historia(db, ticket.id_activo, "MANTENIMIENTO_CORRECTIVO", desc_cierre, usuario)
    
    
    db.commit()
    
    return RedirectResponse("/gestion-novedades?msg=Ticket resuelto exitosamente&tipo=success", status_code=303)
